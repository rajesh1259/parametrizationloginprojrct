# from selenium import webdriver
# import time
#
#
# options=webdriver.ChromeOptions()
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.common.by import By
# import openpyxl
# #excel_path='/Volumes/MAC - Data/excel/newlogin.xlsx'
#
#
# def my_excel(file_path,sheet_name):
#     wk = openpyxl.load_workbook(file_path)
#     sheet = wk[sheet_name]
#     credentials=[]
#     for row in sheet_name.iter_rows(min_row=2,values_only=True):
#         username,password = row
#         credentials.append((username,password))
#
#     return credentials
#
# def my_login(username,password):
#     service = Service(executable_path='/Users/vaibhavlutade/PycharmProjects/pytestproject/newlogin/chromedriver122')
#     driver = webdriver.Chrome(service=service, options=options)
#     driver.get("https://finserv-pa.reposenergy.com/login")
#     time.sleep(5)
#     driver.find_element(By.ID,"Email address / mobile ").send_keys(username)
#     time.sleep(5)
#     driver.find_element(By.ID,"Password").send_keys(password)
#     time.sleep(5)
#     driver.find_element(By.ID,"Log In button").click()
#     time.sleep(5)
#     driver.quit()
#
# file_path='/Volumes/MAC - Data/excel/newlogin.xlsx'
# sheet_name_data='newlogin'
#
# res=my_excel(file_path,sheet_name_data)
#
# for username,password in res:
#     my_login(username,password)
#


from selenium import webdriver
import time
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import openpyxl

options = webdriver.ChromeOptions()  # Correct instantiation of ChromeOptions

def my_excel(file_path, sheet_name):
    wk = openpyxl.load_workbook(file_path)
    #sheet = wk[sheet_name]
    sheet_name=wk.active
    credentials = []
    for row in sheet_name.iter_rows(min_row=2, values_only=True):
        username, password = row
        if username is not None and password is not None:
            credentials.append((username, password))


    return credentials

def my_login(username, password):
    if username is None or password is None:
        print("Skipping row with None values.")
        return
    service = Service(executable_path='/Users/vaibhavlutade/PycharmProjects/pytestproject/newlogin/chromedriver122')
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 10)

    driver.get("https://finserv-pa.reposenergy.com/login")
    time.sleep(5)
    wait.until(EC.presence_of_element_located((By.NAME, 'inputName'))).send_keys(username)
    #driver.find_element(By.NAME, "inputName").send_keys(username)
    driver.find_element(By.ID, "Password").send_keys(password)
    element = wait.until(EC.presence_of_element_located((By.ID, 'Log In button')))
    element.click()

    #driver.find_element(By.ID, "Log In button").click()
    time.sleep(5)
    driver.quit()

file_path = '/Volumes/MAC - Data/excel/datafed1.xlsx'
sheet_name_data = 'datafed1'  # Fixed the variable name

res = my_excel(file_path, sheet_name_data)  # Fixed the variable name
print(res)


for username, password in res:
        my_login(username, password)
